import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook
import player_data  # Import the player data module

# Function to record yellows
def record_yellows():
    try:
        # Specify the path to the original Excel file for yellows
        yellows_excel_file = "F:\OneDrive\TubiStats\YellowsTest.xlsx"

        # Load the existing workbook
        workbook = load_workbook(yellows_excel_file)

        # Get the active sheet (you can modify this to select a specific sheet)
        sheet = workbook.active

        # Determine the next match ID based on the last value in the "matchId" column
        last_match_id = sheet.cell(row=sheet.max_row, column=3).value
        if last_match_id is None:
            match_id = 1
        else:
            match_id = last_match_id + 1

        # Get the selected players (those with checkboxes)
        selected_players = [player_data.players[i] for i, var in enumerate(player_vars) if var.get()]

        # Calculate the starting index based on existing data
        existing_data = [cell.value for cell in sheet["A"]]
        if all(value is None for value in existing_data[1:]) or (len(existing_data) == 1 and existing_data[0] is None):
            start_index = 1
        else:
            start_index = max([int(value) for value in existing_data[1:] if value is not None], default=0) + 1

        # Append each selected player as a separate row with the specified index and the determined match ID
        for i, player in enumerate(selected_players, start=start_index):
            player_id = int(player_data.player_id_dict[player])
            sheet.append([i, player_id, match_id])

        # Save the Excel file with the updated data
        workbook.save(yellows_excel_file)

        # Reset the form (clear the checkboxes)
        for var in player_vars:
            var.set(False)

        # Show a success message
        messagebox.showinfo("Success", "Yellows recorded successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

def create_page(notebook):
    yellows_frame = ttk.Frame(notebook)
    notebook.add(yellows_frame, text="Yellows")

    global player_vars
    player_vars = []

    # Create labels and input fields for goal recording
    bold_font = ("", 10, "bold")  # Define a bold font

    # Create checkboxes for player selection (centered)
    for player in player_data.players:
        var = tk.BooleanVar()
        player_vars.append(var)

        label = ttk.Label(yellows_frame, text=player, font=bold_font)
        label.grid(row=player_data.players.index(player), column=0, padx=10, pady=5, sticky=tk.W)

        checkbox = ttk.Checkbutton(yellows_frame, variable=var)
        checkbox.grid(row=player_data.players.index(player), column=1, padx=10, pady=5, sticky=tk.E)

    # Button to record yellows - Yellows Page
    yellows_button = ttk.Button(yellows_frame, text="Record Yellows", command=record_yellows)
    yellows_button.grid(row=len(player_data.players), columnspan=2, padx=10, pady=10)
    yellows_button.configure(width=20)  # Adjust the button width for better appearance

    # Center the button within the frame
    yellows_frame.columnconfigure(0, weight=1)