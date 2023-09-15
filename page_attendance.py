import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook
import player_data  # Import the player data module

def record_attendance():
    try:
        # Specify the path to the original Excel file
        excel_file = "F:\OneDrive\TubiStats\AttendanceTest.xlsx"

        # Load the existing workbook
        workbook = load_workbook(excel_file)

        # Get the active sheet (you can modify this to select a specific sheet)
        sheet = workbook.active

        # Determine the next training ID based on the last value in the "trainingId" column
        last_training_id = sheet.cell(row=sheet.max_row, column=2).value
        if last_training_id is None:
            training_id = 1
        else:
            training_id = last_training_id + 1

        # Get the selected players (those present) and their IDs
        present_players = [int(player_data.player_id_dict[player]) for player in player_data.players if player_vars[player_data.players.index(player)].get()]

        # Calculate the starting index based on existing data
        existing_data = [cell.value for cell in sheet["A"]]
        if all(value is None for value in existing_data[1:]) or (len(existing_data) == 1 and existing_data[0] is None):
            start_index = 1
        else:
            start_index = max([int(value) for value in existing_data[1:] if value is not None], default=0) + 1

        # Append each player's ID as a separate row with the specified index and the determined training ID
        for i, player_id in enumerate(present_players, start=start_index):
            sheet.append([i, training_id, player_id])

        # Save the Excel file with the updated data
        workbook.save(excel_file)

        # Reset the form (clear the checkboxes)
        for var in player_vars:
            var.set(False)

        # Show a success message
        messagebox.showinfo("Success", "Attendance recorded successfully! Training ID: {}".format(training_id))
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

def create_page(notebook):
    attendance_frame = ttk.Frame(notebook)
    notebook.add(attendance_frame, text="Attendance")

    global player_vars
    player_vars = []

    # Create labels and input fields for goal recording
    bold_font = ("", 10, "bold")  # Define a bold font

    # Create checkboxes and labels for player selection (centered)
    for player in player_data.players:
        var = tk.BooleanVar()
        player_vars.append(var)

        frame = ttk.Frame(attendance_frame)
        frame.pack(fill=tk.X, padx=10, pady=5)

        label = ttk.Label(frame, text=player, font=bold_font)
        label.pack(side=tk.LEFT)

        checkbox = ttk.Checkbutton(frame, variable=var)
        checkbox.pack(side=tk.RIGHT)

    # Button to record attendance - Attendance Page
    record_button = ttk.Button(attendance_frame, text="Record Attendance", command=record_attendance)
    record_button.pack(pady=10)

    # Center the button within the frame
    attendance_frame.columnconfigure(0, weight=1)
