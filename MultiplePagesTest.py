import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook

# Define a dictionary to map player names to player IDs
player_id_dict = {
    "Baran": "1",
    "Cem": "2",
    "Emir": "3",
    "Ersan": "4",
    "Ivo": "5",
    "Jochem": "6",
    "Jordy": "7",
    "Justin": "8",
    "Kaan": "9",
    "Kevin": "10",
    "Koray": "11",
    "Lars": "12",
    "Luc": "13",
    "Oseb": "14",
    "Rick": "15",
    "Stefan": "16",
    "Sven": "17",
    "Swen": "18"
}

# List of player names
players = list(player_id_dict.keys())

player_vars = []

# Function to record attendance
def record_attendance():
    try:
        # Specify the path to the original Excel file
        excel_file = "F:\OneDrive\TubiStats\AttendanceTest.xlsx"

        # Load the existing workbook
        workbook = load_workbook(excel_file)

        # Get the active sheet (you can modify this to select a specific sheet)
        sheet = workbook.active

        # Determine the next training ID based on the last value in the "trainingId" column
        last_training_id = sheet.cell(row=sheet.max_row, column=2).value
        if last_training_id is None:
            training_id = 1
        else:
            training_id = last_training_id + 1

        # Get the selected players (those present) and their IDs
        present_players = [int(player_id_dict[players[i]]) for i, var in enumerate(player_vars) if var.get()]

        # Calculate the starting index based on existing data
        existing_data = [cell.value for cell in sheet["A"]]
        if all(value is None for value in existing_data[1:]) or (len(existing_data) == 1 and existing_data[0] is None):
            start_index = 1
        else:
            start_index = max([int(value) for value in existing_data[1:] if value is not None], default=0) + 1

        # Append each player's ID as a separate row with the specified index and the determined training ID
        for i, player_id in enumerate(present_players, start=start_index):
            sheet.append([i, training_id, player_id])

        # Save the Excel file with the updated data
        workbook.save(excel_file)

        # Reset the form (clear the checkboxes)
        for var in player_vars:
            var.set(False)

        # Show a success message
        messagebox.showinfo("Success", "Attendance recorded successfully! Training ID: {}".format(training_id))
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

# Function to record goals
def record_goals():
    try:
        # Specify the path to the original Excel file for goals
        goals_excel_file = "F:\OneDrive\TubiStats\GoalsTest.xlsx"

        # Your code for recording goals goes here
        
        # Show a success message
        messagebox.showinfo("Success", "Goals recorded successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

# Function to record assists
def record_assists():
    try:
        # Specify the path to the original Excel file for assists
        assists_excel_file = "F:\OneDrive\TubiStats\AssistsTest.xlsx"

        # Your code for recording assists goes here
        
        # Show a success message
        messagebox.showinfo("Success", "Assists recorded successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

# Function to record minutes
def record_minutes():
    try:
        # Specify the path to the original Excel file for minutes
        minutes_excel_file = "F:\OneDrive\TubiStats\MinutesTest.xlsx"

        # Your code for recording minutes goes here
        
        # Show a success message
        messagebox.showinfo("Success", "Minutes recorded successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

# Function to record yellows
def record_yellows():
    try:
        # Specify the path to the original Excel file for yellows
        yellows_excel_file = "F:\OneDrive\TubiStats\YellowsTest.xlsx"

        # Your code for recording yellows goes here
        
        # Show a success message
        messagebox.showinfo("Success", "Yellows recorded successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

# Function to record penalties
def record_penalties():
    try:
        # Specify the path to the original Excel file for penalties
        penalties_excel_file = "F:\OneDrive\TubiStats\PenaltiesTest.xlsx"

        # Your code for recording penalties goes here
        
        # Show a success message
        messagebox.showinfo("Success", "Penalties recorded successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

# Create the main application window
root = tk.Tk()
root.title("Player Statistics Tracker")

# Set the window size to 400x500 pixels
root.geometry("400x500")

# Create a notebook (tabbed interface) for multiple pages
notebook = ttk.Notebook(root)
notebook.pack(fill=tk.BOTH, expand=True)

# Attendance Page
attendance_frame = ttk.Frame(notebook)
notebook.add(attendance_frame, text="Attendance")

# Goals Page
goals_frame = ttk.Frame(notebook)
notebook.add(goals_frame, text="Goals")

# Assists Page
assists_frame = ttk.Frame(notebook)
notebook.add(assists_frame, text="Assists")

# Minutes Page
minutes_frame = ttk.Frame(notebook)
notebook.add(minutes_frame, text="Minutes")

# Yellows Page
yellows_frame = ttk.Frame(notebook)
notebook.add(yellows_frame, text="Yellows")

# Penalties Page
penalties_frame = ttk.Frame(notebook)
notebook.add(penalties_frame, text="Penalties")

# Checkboxes for selecting players (18 players) - Attendance Page
for player in players:
    var = tk.BooleanVar()
    player_vars.append(var)
    checkbox = ttk.Checkbutton(attendance_frame, text=player, variable=var)
    checkbox.pack()

# Button to record attendance - Attendance Page
attendance_button = ttk.Button(attendance_frame, text="Record Attendance", command=record_attendance)
attendance_button.pack()

# Button to record goals - Goals Page
goals_button = ttk.Button(goals_frame, text="Record Goals", command=record_goals)
goals_button.pack()

# Button to record assists - Assists Page
assists_button = ttk.Button(assists_frame, text="Record Assists", command=record_assists)
assists_button.pack()

# Button to record minutes - Minutes Page
minutes_button = ttk.Button(minutes_frame, text="Record Minutes", command=record_minutes)
minutes_button.pack()

# Button to record yellows - Yellows Page
yellows_button = ttk.Button(yellows_frame, text="Record Yellows", command=record_yellows)
yellows_button.pack()

# Button to record penalties - Penalties Page
penalties_button = ttk.Button(penalties_frame, text="Record Penalties", command=record_penalties)
penalties_button.pack()

# Start the GUI application
root.mainloop()