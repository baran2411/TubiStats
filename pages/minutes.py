import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook
import players  # Import the player data module

bold_font = ("", 10, "bold")  # Define a bold font


# Function to record minutes and starting eleven status
def record_minutes():
    try:
        # Specify the path to the original Excel file for minutes
        minutes_excel_file = "F:\OneDrive\TubiStats\Minute.xlsx"

        # Load the existing workbook
        workbook = load_workbook(minutes_excel_file)

        # Get the active sheet (you can modify this to select a specific sheet)
        sheet = workbook.active

        # Determine the next match ID based on the last value in the "matchId" column
        last_match_id = sheet.cell(row=sheet.max_row, column=3).value
        if last_match_id is None:
            match_id = 1
        else:
            match_id = last_match_id + 1

        # Get the player names, minutes, and starting eleven status
        minutes_data = []
        for i, player in enumerate(players.players):
            minutes = minutes_vars[i].get()
            starting_eleven = starting_eleven_vars[i].get()  # Get the checkbox value
            if minutes:
                player_id = int(players.player_id_dict[player])  # Wrap in int() to ensure it's an integer
                minutes = int(minutes)  # Wrap in int() to ensure it's an integer
                basis = True if starting_eleven else False  # Determine the basis value
                minutes_data.append([match_id, player_id, minutes, basis])

        # Calculate the starting index based on existing data
        existing_data = [cell.value for cell in sheet["A"]]
        if all(value is None for value in existing_data[1:]) or (len(existing_data) == 1 and existing_data[0] is None):
            start_index = 1
        else:
            start_index = max([int(value) for value in existing_data[1:] if value is not None], default=0) + 1

        # Append each player's data as a separate row with the specified index and the determined match ID
        for i, (match_id, player_id, minutes, basis) in enumerate(minutes_data, start=start_index):
            sheet.append([i, player_id, match_id, minutes, basis])

        # Save the Excel file with the updated data
        workbook.save(minutes_excel_file)

        # Reset the input fields and checkboxes
        for minutes_var, checkbox_var in zip(minutes_vars, starting_eleven_vars):
            minutes_var.set("")
            checkbox_var.set(False)

        # Show a success message
        messagebox.showinfo("Success", "Minutes and starting eleven status recorded successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

def create_page(notebook):
    minutes_frame = ttk.Frame(notebook)
    notebook.add(minutes_frame, text="Minutes")

    global minutes_vars
    global starting_eleven_vars  # Add a global list for checkbox variables
    minutes_vars = []
    starting_eleven_vars = []

    # Create labels, input fields, and checkboxes for minute recording and starting eleven status
    for player in players.players:
        label = ttk.Label(minutes_frame, text=player, font=bold_font)
        label.grid(column=0, row=players.players.index(player), padx=10, pady=5, sticky=tk.W)

        var = tk.StringVar()
        minutes_vars.append(var)

        minutes_entry = ttk.Entry(minutes_frame, textvariable=var)
        minutes_entry.grid(column=1, row=players.players.index(player), padx=10, pady=5, sticky=tk.E)

        # Create a checkbox variable and checkbox
        checkbox_var = tk.BooleanVar()
        starting_eleven_vars.append(checkbox_var)
        checkbox = ttk.Checkbutton(minutes_frame, variable=checkbox_var, text="11")
        checkbox.grid(column=2, row=players.players.index(player), padx=10, pady=5, sticky=tk.W)

    # Button to record minutes and starting eleven status - Minutes Page
    minutes_button = ttk.Button(minutes_frame, text="Confirm", command=record_minutes)
    minutes_button.grid(column=0, row=len(players.players), columnspan=3, padx=10, pady=10)
    minutes_button.configure(width=20)  # Adjust the button width for better appearance

    # Center the button within the frame
    minutes_frame.columnconfigure(0, weight=1)