import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook
import players  # Import the player data module

# Function to record assists
def record_assists():
    try:
        # Specify the path to the original Excel file for assists
        assists_excel_file = "F:\OneDrive\TubiStats\Assist.xlsx"

        # Load the existing workbook
        workbook = load_workbook(assists_excel_file)

        # Get the active sheet (you can modify this to select a specific sheet)
        sheet = workbook.active

        # Determine the next match ID based on the last value in the "matchId" column
        last_match_id = sheet.cell(row=sheet.max_row, column=3).value
        if last_match_id is None:
            match_id = 1
        else:
            match_id = last_match_id + 1

        # Get the player names and the corresponding assist amounts
        assist_data = []
        for i, player in enumerate(players.players):
            assist_amount = assist_vars[i].get()
            if assist_amount:
                player_id = int(players.player_id_dict[player])  # Wrap in int() to ensure it's an integer
                assist_amount = int(assist_amount)  # Wrap in int() to ensure it's an integer
                assist_data.append([match_id, player_id, assist_amount])

        # Calculate the starting index based on existing data
        existing_data = [cell.value for cell in sheet["A"]]
        if all(value is None for value in existing_data[1:]) or (len(existing_data) == 1 and existing_data[0] is None):
            start_index = 1
        else:
            start_index = max([int(value) for value in existing_data[1:] if value is not None], default=0) + 1

        # Append each player's data as a separate row with the specified index and the determined match ID
        for i, (match_id, player_id, assist_amount) in enumerate(assist_data, start=start_index):
            sheet.append([i, player_id, match_id, assist_amount])

        # Save the Excel file with the updated data
        workbook.save(assists_excel_file)

        # Reset the input fields
        for assist_var in assist_vars:
            assist_var.set("")

        # Show a success message
        messagebox.showinfo("Success", "Assists recorded successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

def create_page(notebook):
    assists_frame = ttk.Frame(notebook)
    notebook.add(assists_frame, text="Assist")

    global assist_vars
    assist_vars = []

    # Create labels and input fields for goal recording
    bold_font = ("", 10, "bold")  # Define a bold font

    # Create labels and input fields for assist recording
    for player in players.players:
        label = ttk.Label(assists_frame, text=player, font=bold_font)
        label.grid(column=0, row=players.players.index(player), padx=10, pady=5, sticky=tk.W)

        var = tk.StringVar()
        assist_vars.append(var)

        assist_entry = ttk.Entry(assists_frame, textvariable=var)
        assist_entry.grid(column=1, row=players.players.index(player), padx=10, pady=5, sticky=tk.E)

    # Button to record assists - Assists Page
    assists_button = ttk.Button(assists_frame, text="Confirm", command=record_assists)
    assists_button.grid(column=0, row=len(players.players), columnspan=2, padx=10, pady=10)
    assists_button.configure(width=20)  # Adjust the button width for better appearance

    # Center the button within the frame
    assists_frame.columnconfigure(0, weight=1)