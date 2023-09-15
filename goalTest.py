import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook

# Define a dictionary to map player names to player IDs
player_id_dict = {
    "Baran": "1",
    "Cem": "2",
    "Emir": "3",
    "Ersan": "4",
    "Ivo": "5",
    "Jochem": "6",
    "Jordy": "7",
    "Justin": "8",
    "Kaan": "9",
    "Kevin": "10",
    "Koray": "11",
    "Lars": "12",
    "Luc": "13",
    "Oseb": "14",
    "Rick": "15",
    "Stefan": "16",
    "Sven": "17",
    "Swen": "18"
}

# List of player names
players = list(player_id_dict.keys())

player_vars = []

# Function to record attendance
def record_attendance():
    try:
        # Specify the path to the original Excel file
        excel_file = "path_to_your_existing_excel_file.xlsx"

        # Load the existing workbook
        workbook = load_workbook(excel_file)

        # Get the active sheet (you can modify this to select a specific sheet)
        sheet = workbook.active

        # Determine the next training ID based on the last value in the "trainingId" column
        last_training_id = sheet.cell(row=sheet.max_row, column=2).value
        if last_training_id is None:
            training_id = 1
        else:
            training_id = last_training_id + 1

        # Get the selected players (those present) and their IDs
        present_players = [int(player_id_dict[players[i]]) for i, var in enumerate(player_vars) if var.get()]

        # Calculate the starting index based on existing data
        existing_data = [cell.value for cell in sheet["A"]]
        if all(value is None for value in existing_data[1:]) or (len(existing_data) == 1 and existing_data[0] is None):
            start_index = 1
        else:
            start_index = max([int(value) for value in existing_data[1:] if value is not None], default=0) + 1

        # Append each player's ID as a separate row with the specified index and the determined training ID
        for i, player_id in enumerate(present_players, start=start_index):
            sheet.append([i, training_id, player_id])

        # Save the Excel file with the updated data
        workbook.save(excel_file)

        # Reset the form (clear the checkboxes)
        for var in player_vars:
            var.set(False)

        # Show a success message
        messagebox.showinfo("Success", "Attendance recorded successfully! Training ID: {}".format(training_id))
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

# Create the main application window
root = tk.Tk()
root.title("Training Attendance Form")

# Set the window size to 300x500 pixels
root.geometry("300x500")

# Checkboxes for selecting players (18 players)
for player in players:
    var = tk.BooleanVar()
    player_vars.append(var)
    checkbox = ttk.Checkbutton(root, text=player, variable=var)
    checkbox.pack()

# Button to record attendance
record_button = ttk.Button(root, text="Record Attendance", command=record_attendance)
record_button.pack()

# Start the GUI application
root.mainloop()